"""
Script to analyze and document custom roles in Flight Control environment.

This script:
1. Lists all custom roles in the Parent CID with detailed permissions
2. Checks which Child CIDs already have these roles
3. Compares permissions between Parent and Child (drift detection)
4. Generates detailed reports including Excel format for manual replication
5. Validates post-replication to ensure correctness

Usage:
    # Analyze and generate reports
    python analyze_roles.py --config config/credentials.json

    # Validate after manual replication
    python analyze_roles.py --config config/credentials.json --validate reports/role_analysis_TIMESTAMP.json
"""

import argparse
import sys
import os
import json
from pathlib import Path
from typing import List, Dict, Any, Set
from datetime import datetime

# Fix Windows console encoding
if sys.platform == 'win32':
    try:
        sys.stdout.reconfigure(encoding='utf-8')
    except:
        pass

# Add parent directory to path for imports
sys.path.insert(0, str(Path(__file__).parent.parent))

from falconpy import FlightControl, UserManagement, APIHarnessV2
from utils.common import check_response, extract_resources
from utils.auth import get_credentials_smart
from utils.formatting import (
    print_header, print_section, print_success, print_error, print_warning, print_info,
    print_progress, print_table, print_summary_box, print_status_indicator,
    print_coverage_bar, print_role_item, print_child_item, create_summary_table,
    print_action_items, print_credentials_source, Colors
)

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False
    print_warning("openpyxl not installed. Excel export will be disabled. Install with: pip install openpyxl")


def compare_permissions(parent_perms: List[str], child_perms: List[str]) -> Dict[str, Any]:
    """
    Compare permissions between Parent and Child role.

    Returns:
        Dict with 'missing', 'extra', 'matching' lists and 'drift_detected' bool
    """
    parent_set = set(parent_perms)
    child_set = set(child_perms)

    missing = sorted(list(parent_set - child_set))  # In Parent but not in Child
    extra = sorted(list(child_set - parent_set))     # In Child but not in Parent
    matching = sorted(list(parent_set & child_set))  # In both

    return {
        "missing": missing,
        "extra": extra,
        "matching": matching,
        "drift_detected": len(missing) > 0 or len(extra) > 0,
        "total_parent": len(parent_perms),
        "total_child": len(child_perms),
        "match_percentage": (len(matching) / len(parent_perms) * 100) if len(parent_perms) > 0 else 0
    }


def is_custom_role(role: Dict[str, Any]) -> bool:
    """
    Determine if a role is custom based on its ID format.
    Custom roles have UUID format IDs, built-in roles have text IDs.
    """
    role_id = role.get('id', '')
    return len(role_id) == 32 and all(c in '0123456789abcdef' for c in role_id.lower())


def get_role_permissions(api_harness: APIHarnessV2, role_id: str, cid: str = None) -> List[str]:
    """
    Get detailed permissions for a role using the low-level API.

    Args:
        api_harness: APIHarnessV2 instance
        role_id: Role ID
        cid: Optional CID

    Returns:
        List of permission strings
    """
    params = {"ids": [role_id]}
    if cid:
        params["cid"] = cid

    response = api_harness.command("entitiesRolesV1", **params)

    if response.get('status_code') == 200:
        resources = response.get('body', {}).get('resources', [])
        if resources:
            role_data = resources[0]
            # Try to get permissions from various possible fields
            permissions = role_data.get('permissions', [])
            if not permissions:
                # Sometimes permissions are in grant_scopes or role_scopes
                permissions = role_data.get('grant_scopes', [])
            if not permissions:
                permissions = role_data.get('role_scopes', [])
            return permissions

    return []


def select_custom_roles_to_analyze(custom_roles: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    """
    Allow user to interactively select which custom roles to analyze.
    """
    print_header("SELECT CUSTOM ROLES TO ANALYZE", width=80)

    if not custom_roles:
        print_warning("No custom roles found.")
        return []

    # Display roles
    print_info(f"Found {len(custom_roles)} custom role(s):")
    print()

    for i, role in enumerate(custom_roles, 1):
        role_name = role.get('display_name', 'N/A')
        role_desc = role.get('description', 'No description')
        role_id = role.get('id')
        print_role_item(i, role_name, role_desc, role_id)

    print_section("", char="-", width=80)
    print()

    # Interactive selection
    print(f"{Colors.INFO}Selection options:{Colors.RESET}")
    print("  • Enter role numbers separated by commas (e.g., 1,3,5)")
    print("  • Enter 'all' to select all custom roles")
    print("  • Enter 'q' to quit")
    print()

    while True:
        selection = input(f"{Colors.HIGHLIGHT}Select roles to analyze:{Colors.RESET} ").strip()

        if selection.lower() == 'q':
            print_warning("Cancelled by user.")
            return []

        if selection.lower() == 'all':
            print_success(f"Selected all {len(custom_roles)} custom roles")
            return custom_roles

        # Parse comma-separated numbers
        try:
            indices = [int(x.strip()) for x in selection.split(',')]
            selected_roles = []

            for idx in indices:
                if 1 <= idx <= len(custom_roles):
                    selected_roles.append(custom_roles[idx - 1])
                else:
                    print_warning(f"Invalid selection {idx} (must be between 1 and {len(custom_roles)})")

            if selected_roles:
                print()
                print_success(f"Selected {len(selected_roles)} role(s):")
                for role in selected_roles:
                    print(f"  • {role.get('display_name')}")
                return selected_roles
            else:
                print_error("No valid roles selected. Please try again.")

        except ValueError:
            print_error("Invalid input. Please enter numbers separated by commas, 'all', or 'q'.")
            continue


def select_children_to_check(children: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    """
    Allow user to interactively select which Child CIDs to check.
    """
    print_header("SELECT CHILD CIDs TO CHECK", width=80)

    if not children:
        print_warning("No child CIDs found.")
        return []

    # Display children
    print_info(f"Found {len(children)} Child CID(s):")
    print()

    for i, child in enumerate(children, 1):
        child_name = child.get('name', 'Unknown')
        child_cid = child.get('child_cid', 'N/A')
        print_child_item(i, child_name, child_cid)

    print_section("", char="-", width=80)
    print()

    # Interactive selection
    print(f"{Colors.INFO}Selection options:{Colors.RESET}")
    print("  • Enter child numbers separated by commas (e.g., 1,2,4)")
    print("  • Enter 'all' to select all children")
    print("  • Enter 'q' to quit")
    print()

    while True:
        selection = input(f"{Colors.HIGHLIGHT}Select children to check:{Colors.RESET} ").strip()

        if selection.lower() == 'q':
            print_warning("Cancelled by user.")
            return []

        if selection.lower() == 'all':
            print_success(f"Selected all {len(children)} children")
            return children

        # Parse comma-separated numbers
        try:
            indices = [int(x.strip()) for x in selection.split(',')]
            selected_children = []

            for idx in indices:
                if 1 <= idx <= len(children):
                    selected_children.append(children[idx - 1])
                else:
                    print_warning(f"Invalid selection {idx} (must be between 1 and {len(children)})")

            if selected_children:
                print()
                print_success(f"Selected {len(selected_children)} child(ren):")
                for child in selected_children:
                    print(f"  • {child.get('name')} (CID: {child.get('child_cid')})")
                return selected_children
            else:
                print_error("No valid children selected. Please try again.")

        except ValueError:
            print_error("Invalid input. Please enter numbers separated by commas, 'all', or 'q'.")
            continue


def get_all_custom_roles(user_mgmt: UserManagement, api_harness: APIHarnessV2, interactive: bool = False) -> List[Dict[str, Any]]:
    """
    Get all custom roles from Parent CID with detailed information.
    """
    print_header("ANALYZING CUSTOM ROLES IN PARENT CID", width=80)

    # Query all roles
    print_info("Querying roles from Parent CID...")
    query_response = user_mgmt.query_roles()

    if not check_response(query_response, "Query roles"):
        return []

    role_ids = extract_resources(query_response)

    if not role_ids:
        print_warning("No roles found")
        return []

    # Get detailed role information
    roles_response = user_mgmt.GetRoles(ids=role_ids)

    if not check_response(roles_response, "Get role details"):
        return []

    all_roles = extract_resources(roles_response)
    custom_roles = []

    print_info(f"Found {len(all_roles)} total roles")
    print_info("Filtering custom roles...")

    for role in all_roles:
        if is_custom_role(role):
            custom_roles.append(role)

    print_success(f"Found {len(custom_roles)} custom role(s)")

    if not custom_roles:
        return []

    # Interactive selection if requested
    if interactive:
        custom_roles = select_custom_roles_to_analyze(custom_roles)
        if not custom_roles:
            return []

    # Now get permissions for selected roles
    print()
    print_info("Retrieving permissions for selected roles...")
    print()

    for i, role in enumerate(custom_roles, 1):
        role_id = role.get('id')
        role_name = role.get('display_name')

        # Progress indicator
        print_progress(i, len(custom_roles), prefix="Analyzing roles", suffix=f"({role_name[:30]}...)")

        # Get permissions for this role
        permissions = get_role_permissions(api_harness, role_id)
        role['permissions'] = permissions

    print()
    print_success("Role analysis complete!")

    return custom_roles


def get_all_children(flight_control: FlightControl, interactive: bool = False) -> List[Dict[str, Any]]:
    """
    Get all child CIDs from the parent.
    """
    print_header("DISCOVERING CHILD CIDs", width=80)

    query_response = flight_control.queryChildren()

    if not check_response(query_response, "Query children"):
        return []

    child_cids = extract_resources(query_response)

    if not child_cids:
        print_warning("No child CIDs found")
        return []

    details_response = flight_control.getChildren(ids=child_cids)

    if not check_response(details_response, "Get child details"):
        return []

    children = extract_resources(details_response)

    print_success(f"Found {len(children)} child CID(s)")

    if not interactive:
        # Non-interactive: just show the list
        print()
        for child in children:
            print(f"  • {child.get('name')} (CID: {child.get('child_cid')})")
        return children

    # Interactive: let user select
    return select_children_to_check(children)


def check_role_in_child(user_mgmt: UserManagement, api_harness: APIHarnessV2, role_name: str, child_cid: str, parent_role: Dict[str, Any]) -> Dict[str, Any]:
    """
    Check if a role exists in a child CID and compare permissions.

    Returns:
        Dict with 'exists' (bool), 'role_data' (if exists), and 'permissions_comparison' (if exists)
    """
    # Query roles in child
    query_response = user_mgmt.query_roles(cid=child_cid)

    if not check_response(query_response, f"Query roles in child {child_cid}", verbose=False):
        return {"exists": False, "error": "Failed to query roles"}

    role_ids = extract_resources(query_response)

    if not role_ids:
        return {"exists": False}

    # Get role details
    roles_response = user_mgmt.GetRoles(ids=role_ids, cid=child_cid)

    if not check_response(roles_response, f"Get role details in child {child_cid}", verbose=False):
        return {"exists": False, "error": "Failed to get role details"}

    roles = extract_resources(roles_response)

    # Find matching role
    for role in roles:
        if role.get('display_name') == role_name:
            # Get detailed permissions for child role
            child_permissions = get_role_permissions(api_harness, role.get('id'), child_cid)
            parent_permissions = parent_role.get('permissions', [])

            # Compare permissions
            comparison = compare_permissions(parent_permissions, child_permissions)

            return {
                "exists": True,
                "role_data": role,
                "permissions_comparison": comparison
            }

    return {"exists": False}


def analyze_role_coverage(user_mgmt: UserManagement, api_harness: APIHarnessV2, custom_roles: List[Dict[str, Any]], children: List[Dict[str, Any]]) -> Dict[str, Any]:
    """
    Analyze which children have which custom roles and compare permissions.
    """
    print_header("ANALYZING ROLE COVERAGE ACROSS CHILDREN", width=80)

    coverage = {}
    total_checks = len(custom_roles) * len(children)
    current_check = 0

    for role in custom_roles:
        role_name = role.get('display_name')
        role_id = role.get('id')

        coverage[role_name] = {
            "role_id": role_id,
            "parent_role": role,
            "children_status": {}
        }

        for child in children:
            child_cid = child.get('child_cid')
            child_name = child.get('name')

            current_check += 1
            print_progress(
                current_check,
                total_checks,
                prefix="Checking coverage",
                suffix=f"({role_name[:20]} in {child_name[:20]})"
            )

            result = check_role_in_child(user_mgmt, api_harness, role_name, child_cid, role)

            coverage[role_name]["children_status"][child_cid] = {
                "name": child_name,
                "exists": result.get("exists", False),
                "role_data": result.get("role_data"),
                "error": result.get("error"),
                "permissions_comparison": result.get("permissions_comparison")
            }

    print()
    print_success("Coverage analysis complete!")

    # Print detailed results with drift detection
    print()
    print_section("DETAILED RESULTS", char="-", width=80)
    print()

    for role_name, data in coverage.items():
        children_status = data.get('children_status', {})
        exists_count = sum(1 for status in children_status.values() if status['exists'])
        drift_count = sum(1 for status in children_status.values()
                         if status.get('exists') and status.get('permissions_comparison', {}).get('drift_detected'))

        print(f"\n{Colors.HIGHLIGHT}▶ {role_name}{Colors.RESET}")
        print_coverage_bar("  Coverage", exists_count, len(children), width=40)

        if drift_count > 0:
            print(f"  {Colors.WARNING}⚠ Configuration drift detected in {drift_count} child(ren){Colors.RESET}")

        print()

        for cid, status in children_status.items():
            child_name = status['name']
            exists = status['exists']

            status_text = f"    {child_name}"

            if exists:
                comparison = status.get('permissions_comparison', {})
                if comparison and comparison.get('drift_detected'):
                    missing = len(comparison.get('missing', []))
                    extra = len(comparison.get('extra', []))
                    match_pct = comparison.get('match_percentage', 0)

                    drift_info = f" ({match_pct:.0f}% match"
                    if missing > 0:
                        drift_info += f", {missing} missing"
                    if extra > 0:
                        drift_info += f", {extra} extra"
                    drift_info += ")"

                    print(f"{Colors.WARNING}    ⚠ {child_name}{drift_info}{Colors.RESET}")
                else:
                    print_status_indicator(status_text, True)
            else:
                print_status_indicator(status_text, False)

    return coverage


def generate_excel_report(custom_roles: List[Dict[str, Any]], children: List[Dict[str, Any]], coverage: Dict[str, Any], output_dir: str = ".") -> str:
    """
    Generate Excel workbook with detailed replication guide.

    Returns:
        Path to the generated Excel file
    """
    if not EXCEL_AVAILABLE:
        print_warning("Excel export skipped (openpyxl not installed)")
        return None

    output_path = Path(output_dir)
    output_path.mkdir(exist_ok=True)

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    excel_file = output_path / f"role_replication_guide_{timestamp}.xlsx"

    wb = Workbook()

    # Styles
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=12)
    subheader_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    subheader_font = Font(bold=True, size=11)
    warning_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    error_fill = PatternFill(start_color="F4CCCC", end_color="F4CCCC", fill_type="solid")
    success_fill = PatternFill(start_color="D9EAD3", end_color="D9EAD3", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Sheet 1: Summary
    ws_summary = wb.active
    ws_summary.title = "Summary"

    ws_summary['A1'] = "Custom Roles Replication Summary"
    ws_summary['A1'].font = Font(size=16, bold=True)
    ws_summary.merge_cells('A1:E1')

    ws_summary['A3'] = "Generated:"
    ws_summary['B3'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')

    ws_summary['A4'] = "Total Custom Roles:"
    ws_summary['B4'] = len(custom_roles)

    ws_summary['A5'] = "Total Child CIDs:"
    ws_summary['B5'] = len(children)

    # Coverage matrix
    ws_summary['A7'] = "Coverage Matrix"
    ws_summary['A7'].font = subheader_font
    ws_summary['A7'].fill = subheader_fill

    # Headers
    row = 8
    ws_summary[f'A{row}'] = "Role Name"
    ws_summary[f'A{row}'].font = header_font
    ws_summary[f'A{row}'].fill = header_fill

    for col_idx, child in enumerate(children, start=2):
        cell = ws_summary.cell(row=row, column=col_idx)
        cell.value = child.get('name')
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', wrap_text=True)

    # Data rows
    for role_name, data in coverage.items():
        row += 1
        ws_summary[f'A{row}'] = role_name
        ws_summary[f'A{row}'].font = Font(bold=True)

        children_status = data.get('children_status', {})

        for col_idx, child in enumerate(children, start=2):
            child_cid = child.get('child_cid')
            status = children_status.get(child_cid, {})
            cell = ws_summary.cell(row=row, column=col_idx)

            if status.get('exists'):
                comparison = status.get('permissions_comparison', {})
                if comparison and comparison.get('drift_detected'):
                    match_pct = comparison.get('match_percentage', 0)
                    cell.value = f"⚠ {match_pct:.0f}%"
                    cell.fill = warning_fill
                else:
                    cell.value = "✓"
                    cell.fill = success_fill
            else:
                cell.value = "✗"
                cell.fill = error_fill

            cell.alignment = Alignment(horizontal='center')
            cell.border = thin_border

    # Auto-adjust column widths
    ws_summary.column_dimensions['A'].width = 40
    for col_idx in range(2, 2 + len(children)):
        ws_summary.column_dimensions[chr(64 + col_idx)].width = 20

    # Sheet 2: Child CIDs
    ws_children = wb.create_sheet("Child CIDs")
    ws_children['A1'] = "Child CID"
    ws_children['B1'] = "CID"
    ws_children['A1'].font = header_font
    ws_children['A1'].fill = header_fill
    ws_children['B1'].font = header_font
    ws_children['B1'].fill = header_fill

    for idx, child in enumerate(children, start=2):
        ws_children[f'A{idx}'] = child.get('name')
        ws_children[f'B{idx}'] = child.get('child_cid')

    ws_children.column_dimensions['A'].width = 40
    ws_children.column_dimensions['B'].width = 35

    # Sheet for each role with permissions
    for role in custom_roles:
        role_name = role.get('display_name', 'Unknown')
        # Sanitize sheet name (Excel limit: 31 chars, no special chars)
        sheet_name = role_name[:31].replace('/', '-').replace('\\', '-').replace(':', '-')

        ws_role = wb.create_sheet(sheet_name)

        # Role header
        ws_role['A1'] = "Role Name:"
        ws_role['B1'] = role_name
        ws_role['A1'].font = Font(bold=True)
        ws_role['B1'].font = Font(size=12)

        ws_role['A2'] = "Description:"
        ws_role['B2'] = role.get('description', 'No description')

        ws_role['A3'] = "Role ID:"
        ws_role['B3'] = role.get('id')

        # Coverage status
        role_coverage = coverage.get(role_name, {})
        children_status = role_coverage.get('children_status', {})

        ws_role['A5'] = "Replication Status by Child CID"
        ws_role['A5'].font = subheader_font
        ws_role['A5'].fill = subheader_fill

        row = 6
        ws_role[f'A{row}'] = "Child CID"
        ws_role[f'B{row}'] = "Status"
        ws_role[f'C{row}'] = "Permissions Match"
        ws_role[f'D{row}'] = "Missing Permissions"
        ws_role[f'E{row}'] = "Extra Permissions"

        for col in ['A', 'B', 'C', 'D', 'E']:
            ws_role[f'{col}{row}'].font = header_font
            ws_role[f'{col}{row}'].fill = header_fill

        for child in children:
            row += 1
            child_cid = child.get('child_cid')
            child_name = child.get('name')
            status = children_status.get(child_cid, {})

            ws_role[f'A{row}'] = child_name

            if status.get('exists'):
                ws_role[f'B{row}'] = "✓ Exists"
                ws_role[f'B{row}'].fill = success_fill

                comparison = status.get('permissions_comparison', {})
                if comparison:
                    match_pct = comparison.get('match_percentage', 0)
                    ws_role[f'C{row}'] = f"{match_pct:.0f}%"

                    missing_count = len(comparison.get('missing', []))
                    extra_count = len(comparison.get('extra', []))

                    ws_role[f'D{row}'] = missing_count
                    ws_role[f'E{row}'] = extra_count

                    if comparison.get('drift_detected'):
                        ws_role[f'C{row}'].fill = warning_fill
                        if missing_count > 0:
                            ws_role[f'D{row}'].fill = error_fill
                        if extra_count > 0:
                            ws_role[f'E{row}'].fill = warning_fill
                    else:
                        ws_role[f'C{row}'].fill = success_fill
            else:
                ws_role[f'B{row}'] = "✗ Missing"
                ws_role[f'B{row}'].fill = error_fill

        # Permissions list
        permissions = role.get('permissions', [])
        perm_start_row = row + 3

        ws_role[f'A{perm_start_row}'] = f"Permissions ({len(permissions)} total)"
        ws_role[f'A{perm_start_row}'].font = subheader_font
        ws_role[f'A{perm_start_row}'].fill = subheader_fill

        ws_role[f'A{perm_start_row + 1}'] = "☐"
        ws_role[f'B{perm_start_row + 1}'] = "Permission"
        ws_role[f'A{perm_start_row + 1}'].font = header_font
        ws_role[f'A{perm_start_row + 1}'].fill = header_fill
        ws_role[f'B{perm_start_row + 1}'].font = header_font
        ws_role[f'B{perm_start_row + 1}'].fill = header_fill

        for idx, perm in enumerate(permissions, start=perm_start_row + 2):
            ws_role[f'A{idx}'] = "☐"
            ws_role[f'B{idx}'] = perm
            ws_role[f'A{idx}'].alignment = Alignment(horizontal='center')

        # Column widths
        ws_role.column_dimensions['A'].width = 5
        ws_role.column_dimensions['B'].width = 60
        ws_role.column_dimensions['C'].width = 18
        ws_role.column_dimensions['D'].width = 20
        ws_role.column_dimensions['E'].width = 20

    # Save
    wb.save(excel_file)
    return str(excel_file)


def generate_report(custom_roles: List[Dict[str, Any]], children: List[Dict[str, Any]], coverage: Dict[str, Any], output_dir: str = "."):
    """
    Generate detailed reports in multiple formats.
    """
    print("\n" + "="*80)
    print("GENERATING REPORTS")
    print("="*80)
    print()

    output_path = Path(output_dir)
    output_path.mkdir(exist_ok=True)

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")

    # 1. JSON Report (detailed data)
    json_file = output_path / f"role_analysis_{timestamp}.json"
    report_data = {
        "timestamp": datetime.now().isoformat(),
        "parent_cid": custom_roles[0].get('cid') if custom_roles else None,
        "custom_roles": custom_roles,
        "children": children,
        "coverage": coverage
    }

    with open(json_file, 'w', encoding='utf-8') as f:
        json.dump(report_data, f, indent=2)

    print(f"✓ JSON report saved: {json_file}")

    # 2. Text Report (human readable with drift detection)
    txt_file = output_path / f"role_analysis_{timestamp}.txt"

    with open(txt_file, 'w', encoding='utf-8') as f:
        f.write("="*80 + "\n")
        f.write("CROWDSTRIKE FALCON - CUSTOM ROLES ANALYSIS REPORT\n")
        f.write("="*80 + "\n")
        f.write(f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
        f.write(f"Parent CID: {custom_roles[0].get('cid') if custom_roles else 'N/A'}\n")
        f.write(f"Custom Roles: {len(custom_roles)}\n")
        f.write(f"Child CIDs: {len(children)}\n")
        f.write("="*80 + "\n\n")

        # Summary by role
        f.write("SUMMARY BY ROLE\n")
        f.write("-"*80 + "\n\n")

        for role_name, data in coverage.items():
            f.write(f"Role: {role_name}\n")
            f.write(f"  ID: {data['role_id']}\n")
            f.write(f"  Description: {data['parent_role'].get('description', 'N/A')}\n")

            exists_count = sum(1 for status in data['children_status'].values() if status['exists'])
            missing_count = len(children) - exists_count
            drift_count = sum(1 for status in data['children_status'].values()
                            if status.get('exists') and status.get('permissions_comparison', {}).get('drift_detected'))

            f.write(f"  Status: {exists_count}/{len(children)} children have this role\n")

            if drift_count > 0:
                f.write(f"  ⚠ Configuration drift detected in {drift_count} child(ren)\n")

            if missing_count > 0:
                f.write(f"  Missing in:\n")
                for cid, status in data['children_status'].items():
                    if not status['exists']:
                        f.write(f"    - {status['name']} (CID: {cid})\n")

            # Drift details
            if drift_count > 0:
                f.write(f"  Drift details:\n")
                for cid, status in data['children_status'].items():
                    if status.get('exists'):
                        comparison = status.get('permissions_comparison', {})
                        if comparison and comparison.get('drift_detected'):
                            child_name = status['name']
                            missing = len(comparison.get('missing', []))
                            extra = len(comparison.get('extra', []))
                            match_pct = comparison.get('match_percentage', 0)

                            f.write(f"    - {child_name}: {match_pct:.0f}% match")
                            if missing > 0:
                                f.write(f", {missing} missing permissions")
                            if extra > 0:
                                f.write(f", {extra} extra permissions")
                            f.write("\n")

            permissions = data['parent_role'].get('permissions', [])
            if permissions:
                f.write(f"  Permissions ({len(permissions)}):\n")
                for perm in permissions[:10]:  # Show first 10
                    f.write(f"    - {perm}\n")
                if len(permissions) > 10:
                    f.write(f"    ... and {len(permissions) - 10} more\n")
            else:
                f.write(f"  Permissions: Unable to retrieve (see JSON report)\n")

            f.write("\n")

        # Summary by child
        f.write("\n" + "="*80 + "\n")
        f.write("SUMMARY BY CHILD CID\n")
        f.write("-"*80 + "\n\n")

        for child in children:
            child_cid = child.get('child_cid')
            child_name = child.get('name')

            f.write(f"Child: {child_name}\n")
            f.write(f"  CID: {child_cid}\n")

            has_roles = []
            missing_roles = []

            for role_name, data in coverage.items():
                status = data['children_status'].get(child_cid, {})
                if status.get('exists'):
                    has_roles.append(role_name)
                else:
                    missing_roles.append(role_name)

            f.write(f"  Has {len(has_roles)}/{len(custom_roles)} custom roles\n")

            if has_roles:
                f.write(f"  Existing roles:\n")
                for role in has_roles:
                    f.write(f"    ✓ {role}\n")

            if missing_roles:
                f.write(f"  Missing roles:\n")
                for role in missing_roles:
                    f.write(f"    ✗ {role}\n")

            f.write("\n")

    print(f"✓ Text report saved: {txt_file}")

    # 3. Manual Replication Guide
    guide_file = output_path / f"manual_replication_guide_{timestamp}.md"

    with open(guide_file, 'w', encoding='utf-8') as f:
        f.write("# Custom Roles Manual Replication Guide\n\n")
        f.write(f"**Generated:** {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}  \n")
        f.write(f"**Parent CID:** `{custom_roles[0].get('cid') if custom_roles else 'N/A'}`  \n\n")

        f.write("## Overview\n\n")
        f.write("This guide provides step-by-step instructions to manually replicate custom roles ")
        f.write("from the Parent CID to Child CIDs in your Flight Control environment.\n\n")

        f.write("## Important Notes\n\n")
        f.write("- Custom roles **cannot be created via API** in CrowdStrike Falcon\n")
        f.write("- Roles must be created manually through the Falcon console\n")
        f.write("- Ensure you have appropriate permissions to manage roles in Child CIDs\n\n")

        f.write("## Custom Roles to Replicate\n\n")

        for i, role in enumerate(custom_roles, 1):
            role_name = role.get('display_name')
            role_desc = role.get('description', 'No description')
            permissions = role.get('permissions', [])

            f.write(f"### {i}. {role_name}\n\n")
            f.write(f"**Description:** {role_desc}  \n")
            f.write(f"**Role ID:** `{role.get('id')}`  \n")
            f.write(f"**Permissions Count:** {len(permissions)}  \n\n")

            # Check which children need this role
            role_coverage = coverage.get(role_name, {})
            children_status = role_coverage.get('children_status', {})

            missing_in = []
            exists_in = []

            for cid, status in children_status.items():
                if status['exists']:
                    exists_in.append(status['name'])
                else:
                    missing_in.append(status['name'])

            if exists_in:
                f.write(f"**✓ Already exists in:** {', '.join(exists_in)}  \n")

            if missing_in:
                f.write(f"**✗ Missing in:** {', '.join(missing_in)}  \n")

            f.write("\n")

            if permissions:
                f.write(f"<details>\n<summary>View all {len(permissions)} permission(s)</summary>\n\n")
                f.write("```\n")
                for perm in permissions:
                    f.write(f"{perm}\n")
                f.write("```\n</details>\n\n")
            else:
                f.write("⚠️ **Note:** Permissions could not be automatically retrieved. ")
                f.write("Please review the role in the Parent CID console.\n\n")

        f.write("## Step-by-Step Replication Instructions\n\n")
        f.write("### For Each Role:\n\n")
        f.write("1. **Login to Falcon Console**\n")
        f.write("   - Navigate to the Parent CID\n")
        f.write("   - Go to: **Support and resources** > **User management** > **Roles**\n\n")

        f.write("2. **Review Role Details**\n")
        f.write("   - Locate the custom role by name\n")
        f.write("   - Click to view its permissions\n")
        f.write("   - Take note of all assigned permissions\n\n")

        f.write("3. **Switch to Child CID**\n")
        f.write("   - Use the CID selector to switch to target Child CID\n")
        f.write("   - Go to: **Support and resources** > **User management** > **Roles**\n\n")

        f.write("4. **Create New Role**\n")
        f.write("   - Click **Create custom role**\n")
        f.write("   - Enter the exact role name from Parent\n")
        f.write("   - Enter the description\n")
        f.write("   - Assign the same permissions as in Parent\n")
        f.write("   - Click **Save**\n\n")

        f.write("5. **Verify**\n")
        f.write("   - Confirm the role appears in the Child CID\n")
        f.write("   - Verify all permissions are correctly assigned\n\n")

        f.write("### Repeat for Each Child CID\n\n")
        f.write("Use the coverage information above to determine which Child CIDs ")
        f.write("need each role.\n\n")

        f.write("## Replication Checklist\n\n")

        for role in custom_roles:
            role_name = role.get('display_name')
            role_coverage = coverage.get(role_name, {})
            children_status = role_coverage.get('children_status', {})

            f.write(f"### {role_name}\n\n")

            for child in children:
                child_name = child.get('name')
                child_cid = child.get('child_cid')
                status = children_status.get(child_cid, {})

                if status.get('exists'):
                    f.write(f"- [x] {child_name}\n")
                else:
                    f.write(f"- [ ] {child_name}\n")

            f.write("\n")

        f.write("## Additional Resources\n\n")
        f.write("- [CrowdStrike Falcon Documentation](https://falcon.crowdstrike.com/documentation)\n")
        f.write("- [User Management Guide](https://falcon.crowdstrike.com/documentation/84/user-management)\n")
        f.write("- For detailed permission data, refer to the JSON report\n\n")

    print(f"✓ Replication guide saved: {guide_file}")

    # 4. Excel Report (interactive replication guide)
    excel_file = generate_excel_report(custom_roles, children, coverage, output_dir)
    if excel_file:
        print(f"✓ Excel guide saved: {excel_file}")

    return {
        "json_report": str(json_file),
        "text_report": str(txt_file),
        "replication_guide": str(guide_file),
        "excel_guide": excel_file
    }


def check_response(response: Dict[str, Any], operation: str, verbose: bool = True) -> bool:
    """Check API response and optionally print errors."""
    status_code = response.get('status_code')

    if status_code in [200, 201, 202, 204]:
        return True

    if verbose:
        print(f"ERROR: {operation} failed (status: {status_code})")
        if 'body' in response and 'errors' in response['body']:
            for error in response['body']['errors']:
                print(f"  {error.get('message', 'Unknown error')}")

    return False


def validate_replication(snapshot_file: str, client_id: str, client_secret: str, base_url: str):
    """
    Validate that roles have been correctly replicated by comparing current state
    with a previous snapshot.
    """
    print_header("VALIDATION MODE - CHECKING REPLICATION STATUS", width=80, color=Colors.INFO)

    # Load snapshot
    snapshot_path = Path(snapshot_file)
    if not snapshot_path.exists():
        print_error(f"Snapshot file not found: {snapshot_file}")
        sys.exit(1)

    print_info(f"Loading snapshot: {snapshot_file}")
    with open(snapshot_path, 'r', encoding='utf-8') as f:
        snapshot = json.load(f)

    snapshot_time = snapshot.get('timestamp', 'Unknown')
    print_success(f"Snapshot loaded (created: {snapshot_time})")
    print()

    # Initialize API clients
    print_info("Authenticating to Falcon API...")
    flight_control = FlightControl(
        client_id=client_id,
        client_secret=client_secret,
        base_url=base_url
    )

    user_mgmt = UserManagement(
        client_id=client_id,
        client_secret=client_secret,
        base_url=base_url
    )

    api_harness = APIHarnessV2(
        client_id=client_id,
        client_secret=client_secret,
        base_url=base_url
    )

    print_success("Authentication successful!")
    print()

    # Get snapshot data
    snapshot_roles = snapshot.get('custom_roles', [])
    snapshot_children = snapshot.get('children', [])
    snapshot_coverage = snapshot.get('coverage', {})

    print_info(f"Snapshot contains {len(snapshot_roles)} role(s) across {len(snapshot_children)} Child CID(s)")
    print()

    # Re-analyze current state
    print_header("RE-ANALYZING CURRENT STATE", width=80)

    # Get current children
    query_response = flight_control.queryChildren()
    child_cids = extract_resources(query_response)
    details_response = flight_control.getChildren(ids=child_cids)
    current_children = extract_resources(details_response)

    # Filter to only children that were in snapshot
    snapshot_cids = {c.get('child_cid') for c in snapshot_children}
    current_children = [c for c in current_children if c.get('child_cid') in snapshot_cids]

    print_success(f"Found {len(current_children)} Child CID(s) from snapshot")
    print()

    # Analyze current coverage
    current_coverage = {}
    validation_results = {}

    for role in snapshot_roles:
        role_name = role.get('display_name')
        role_id = role.get('id')

        print_info(f"Validating: {role_name}")

        current_coverage[role_name] = {
            "role_id": role_id,
            "parent_role": role,
            "children_status": {}
        }

        validation_results[role_name] = {
            "total": len(current_children),
            "passed": 0,
            "failed": 0,
            "drift": 0,
            "details": {}
        }

        for child in current_children:
            child_cid = child.get('child_cid')
            child_name = child.get('name')

            # Get current status
            result = check_role_in_child(user_mgmt, api_harness, role_name, child_cid, role)

            current_coverage[role_name]["children_status"][child_cid] = {
                "name": child_name,
                "exists": result.get("exists", False),
                "role_data": result.get("role_data"),
                "error": result.get("error"),
                "permissions_comparison": result.get("permissions_comparison")
            }

            # Compare with snapshot
            snapshot_status = snapshot_coverage.get(role_name, {}).get('children_status', {}).get(child_cid, {})
            snapshot_existed = snapshot_status.get('exists', False)
            currently_exists = result.get('exists', False)

            validation_details = {
                "child_name": child_name,
                "expected": "exists" if not snapshot_existed else "exists",
                "actual": "exists" if currently_exists else "missing",
                "status": "pass" if currently_exists else "fail"
            }

            if currently_exists:
                comparison = result.get('permissions_comparison', {})
                if comparison:
                    if comparison.get('drift_detected'):
                        validation_details["status"] = "drift"
                        validation_details["drift_info"] = {
                            "missing": len(comparison.get('missing', [])),
                            "extra": len(comparison.get('extra', [])),
                            "match_percentage": comparison.get('match_percentage', 0)
                        }
                        validation_results[role_name]["drift"] += 1
                    else:
                        validation_results[role_name]["passed"] += 1
                else:
                    validation_results[role_name]["passed"] += 1
            else:
                validation_results[role_name]["failed"] += 1

            validation_results[role_name]["details"][child_cid] = validation_details

        print()

    # Display results
    print_header("VALIDATION RESULTS", width=80, color=Colors.SUCCESS)
    print()

    total_checks = 0
    total_passed = 0
    total_failed = 0
    total_drift = 0

    for role_name, results in validation_results.items():
        total_checks += results["total"]
        total_passed += results["passed"]
        total_failed += results["failed"]
        total_drift += results["drift"]

        passed = results["passed"]
        failed = results["failed"]
        drift = results["drift"]
        total = results["total"]

        status_color = Colors.SUCCESS if failed == 0 else Colors.ERROR
        drift_indicator = f" ({drift} with drift)" if drift > 0 else ""

        print(f"{status_color}▶ {role_name}{Colors.RESET}")
        print(f"  {passed}/{total} passed, {failed} failed{drift_indicator}")

        # Show failed or drift details
        for cid, detail in results["details"].items():
            if detail["status"] == "fail":
                print(f"  {Colors.ERROR}  ✗ {detail['child_name']}: Role still missing{Colors.RESET}")
            elif detail["status"] == "drift":
                drift_info = detail.get("drift_info", {})
                match_pct = drift_info.get("match_percentage", 0)
                missing = drift_info.get("missing", 0)
                extra = drift_info.get("extra", 0)
                print(f"  {Colors.WARNING}  ⚠ {detail['child_name']}: {match_pct:.0f}% match ({missing} missing, {extra} extra){Colors.RESET}")

        print()

    # Overall summary
    success_rate = (total_passed / total_checks * 100) if total_checks > 0 else 0

    print_section("", char="=", width=80)
    print_summary_box("OVERALL VALIDATION SUMMARY", {
        "Total checks": total_checks,
        "Passed": f"{total_passed} ({success_rate:.1f}%)",
        "Failed (still missing)": total_failed,
        "Drift detected": total_drift,
        "Validation status": "✓ PASSED" if total_failed == 0 else "✗ INCOMPLETE"
    }, width=80)

    if total_failed > 0:
        print()
        print_warning(f"{total_failed} role(s) still need to be created manually in Child CIDs")

    if total_drift > 0:
        print()
        print_warning(f"{total_drift} role(s) have permission drift and may need adjustment")

    print()

    # Generate validation report
    output_dir = Path(snapshot_file).parent
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    validation_file = output_dir / f"validation_report_{timestamp}.json"

    validation_report = {
        "timestamp": datetime.now().isoformat(),
        "snapshot_file": str(snapshot_file),
        "snapshot_timestamp": snapshot_time,
        "validation_results": validation_results,
        "summary": {
            "total_checks": total_checks,
            "passed": total_passed,
            "failed": total_failed,
            "drift": total_drift,
            "success_rate": success_rate
        }
    }

    with open(validation_file, 'w', encoding='utf-8') as f:
        json.dump(validation_report, f, indent=2)

    print_success(f"Validation report saved: {validation_file}")
    print()

    sys.exit(0 if total_failed == 0 else 1)


def main():
    parser = argparse.ArgumentParser(
        description="Analyze custom roles and generate replication reports for Flight Control",
        formatter_class=argparse.RawDescriptionHelpFormatter
    )

    parser.add_argument(
        "--client-id",
        help="Parent CID API Client ID (can also use FALCON_CLIENT_ID env var)"
    )
    parser.add_argument(
        "--client-secret",
        help="Parent CID API Client Secret (can also use FALCON_CLIENT_SECRET env var)"
    )
    parser.add_argument(
        "--config",
        help="Path to credentials JSON file (highest priority)"
    )
    parser.add_argument(
        "--output-dir",
        default="reports",
        help="Directory to save reports (default: reports/)"
    )
    parser.add_argument(
        "--base-url",
        default="https://api.crowdstrike.com",
        help="Falcon API base URL"
    )
    parser.add_argument(
        "--non-interactive",
        action="store_true",
        help="Run in non-interactive mode (analyze all roles and all children)"
    )
    parser.add_argument(
        "--validate",
        help="Validation mode: compare current state with a previous snapshot JSON file"
    )

    args = parser.parse_args()

    # Get credentials using smart fallback logic
    client_id, client_secret, base_url, source = get_credentials_smart(
        config_path=args.config,
        client_id=args.client_id,
        client_secret=args.client_secret,
        base_url=args.base_url
    )

    if not client_id or not client_secret:
        print("ERROR: No credentials found!")
        print()
        print("Please provide credentials using one of these methods:")
        print("  1. Config file:        --config config/credentials.json")
        print("  2. CLI arguments:      --client-id <ID> --client-secret <SECRET>")
        print("  3. Environment vars:   Set FALCON_CLIENT_ID and FALCON_CLIENT_SECRET")
        print()
        print("PowerShell example:")
        print('  $env:FALCON_CLIENT_ID = "your_client_id"')
        print('  $env:FALCON_CLIENT_SECRET = "your_client_secret"')
        sys.exit(1)

    # Validation mode
    if args.validate:
        validate_replication(args.validate, client_id, client_secret, base_url)
        return  # validate_replication exits on its own

    try:
        print_header("FLIGHT CONTROL - CUSTOM ROLES ANALYZER", width=80, color=Colors.SUCCESS)
        print_credentials_source(source)

        # Initialize API clients
        print_info("Authenticating to Falcon API...")
        flight_control = FlightControl(
            client_id=client_id,
            client_secret=client_secret,
            base_url=base_url
        )

        user_mgmt = UserManagement(
            client_id=client_id,
            client_secret=client_secret,
            base_url=base_url
        )

        api_harness = APIHarnessV2(
            client_id=client_id,
            client_secret=client_secret,
            base_url=base_url
        )

        print_success("Authentication successful!")

        # Determine if interactive mode
        interactive = not args.non_interactive

        if interactive:
            print_section("INTERACTIVE MODE", color=Colors.HIGHLIGHT)
            print_info("You will be able to select which roles and children to analyze.")
            print()
        else:
            print_section("NON-INTERACTIVE MODE", color=Colors.WARNING)
            print_info("Analyzing all roles and all children.")
            print()

        # Step 1: Get custom roles (with optional selection)
        custom_roles = get_all_custom_roles(user_mgmt, api_harness, interactive=interactive)

        if not custom_roles:
            print_warning("No custom roles selected or found in Parent CID")
            sys.exit(0)

        # Step 2: Get children (with optional selection)
        children = get_all_children(flight_control, interactive=interactive)

        if not children:
            print_warning("No child CIDs selected or found")
            sys.exit(0)

        # Step 3: Analyze coverage
        coverage = analyze_role_coverage(user_mgmt, api_harness, custom_roles, children)

        # Step 4: Display summary
        create_summary_table(coverage, children)
        print_action_items(coverage, children)

        # Step 5: Generate reports
        print_section("GENERATING REPORTS", color=Colors.INFO)
        reports = generate_report(custom_roles, children, coverage, args.output_dir)

        # Final summary
        print_summary_box("ANALYSIS COMPLETE", {
            "Custom roles analyzed": len(custom_roles),
            "Child CIDs checked": len(children),
            "Total checks performed": len(custom_roles) * len(children),
            "JSON Report": reports['json_report'],
            "Text Report": reports['text_report'],
            "Replication Guide": reports['replication_guide'],
            "Excel Guide": reports.get('excel_guide', 'Not generated')
        }, width=90)

        print_success("All reports generated successfully!", prefix="✓")
        print()
        print_info("Next steps:")
        print("  1. Review the reports in the output directory")
        print("  2. Use the Excel guide for step-by-step manual replication")
        print("  3. After replication, validate with:")
        print(f"     python analyze_roles.py --config <config> --validate {reports['json_report']}")
        print()

    except KeyboardInterrupt:
        print()
        print_warning("Operation cancelled by user")
        sys.exit(0)
    except Exception as e:
        print()
        print_error(f"Error: {e}")
        import traceback
        traceback.print_exc()
        sys.exit(1)
        print("="*80)
        print("FLIGHT CONTROL - CUSTOM ROLES ANALYZER")
        print("="*80)
        print()
        print(f"Credentials source: {source}")
        print()

        # Initialize API clients
        print("Authenticating...")
        flight_control = FlightControl(
            client_id=client_id,
            client_secret=client_secret,
            base_url=base_url
        )

        user_mgmt = UserManagement(
            client_id=client_id,
            client_secret=client_secret,
            base_url=base_url
        )

        api_harness = APIHarnessV2(
            client_id=client_id,
            client_secret=client_secret,
            base_url=base_url
        )

        print("✓ Authentication successful\n")

        # Determine if interactive mode
        interactive = not args.non_interactive

        if interactive:
            print("="*80)
            print("INTERACTIVE MODE")
            print("="*80)
            print("You will be able to select which roles and children to analyze.")
            print()
        else:
            print("="*80)
            print("NON-INTERACTIVE MODE")
            print("="*80)
            print("Analyzing all roles and all children.")
            print()

        # Step 1: Get custom roles (with optional selection)
        custom_roles = get_all_custom_roles(user_mgmt, api_harness, interactive=interactive)

        if not custom_roles:
            print("\nNo custom roles selected or found in Parent CID")
            sys.exit(0)

        # Step 2: Get children (with optional selection)
        children = get_all_children(flight_control, interactive=interactive)

        if not children:
            print("\nNo child CIDs selected or found")
            sys.exit(0)

        # Step 3: Analyze coverage
        coverage = analyze_role_coverage(user_mgmt, api_harness, custom_roles, children)

        # Step 4: Generate reports
        reports = generate_report(custom_roles, children, coverage, args.output_dir)

        # Final summary
        print("\n" + "="*80)
        print("ANALYSIS COMPLETE")
        print("="*80)
        print(f"\nCustom roles analyzed: {len(custom_roles)}")
        print(f"Child CIDs checked: {len(children)}")
        print(f"\nReports generated:")
        print(f"  - JSON Report: {reports['json_report']}")
        print(f"  - Text Report: {reports['text_report']}")
        print(f"  - Replication Guide: {reports['replication_guide']}")
        print("\n" + "="*80)

    except KeyboardInterrupt:
        print()
        print_warning("Operation cancelled by user")
        sys.exit(0)
    except Exception as e:
        print()
        print_error(f"Error: {e}")
        import traceback
        traceback.print_exc()
        sys.exit(1)


if __name__ == "__main__":
    main()
