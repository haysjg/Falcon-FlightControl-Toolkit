#!/usr/bin/env python3
"""
Export devices, groups, and policies to CSV/Excel for Flight Control CIDs.

This script exports detailed device information including:
- Device details (hostname, OS, etc.)
- Host groups
- Prevention policies (applied vs assigned)
- Response policies (applied vs assigned)
- Sensor update policies (applied vs assigned)

Features:
- Multi-format export (CSV and Excel)
- Device filtering (platform, status, groups, last seen)
- Statistics and anomaly detection
- Interactive CID selection in Flight Control environments
"""

import sys
import os
import csv
import argparse
from datetime import datetime, timedelta
from typing import List, Dict, Any, Optional, Tuple
import json
from collections import Counter

# Add parent directory to path for imports
sys.path.insert(0, os.path.join(os.path.dirname(__file__), '..'))

from falconpy import Hosts, HostGroup, PreventionPolicy, ResponsePolicies, SensorUpdatePolicies, FlightControl
from utils.auth import get_credentials_smart
from utils.formatting import (
    print_header, print_success, print_error, print_info, print_warning,
    print_progress, print_summary_box, print_jg_logo, Colors, print_child_item
)

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
    from openpyxl.utils import get_column_letter
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False

# Fix Windows console encoding for emojis
if sys.platform == 'win32':
    try:
        sys.stdout.reconfigure(encoding='utf-8')
    except:
        pass


class DeviceFilters:
    """Device filtering configuration."""
    def __init__(self, platforms=None, statuses=None, groups=None, stale_days=None):
        self.platforms = [p.lower() for p in platforms] if platforms else None
        self.statuses = [s.lower() for s in statuses] if statuses else None
        self.groups = groups or None
        self.stale_days = stale_days

    def should_include(self, device: Dict[str, Any], host_groups: Dict[str, str]) -> bool:
        """Check if device passes filters."""
        # Platform filter
        if self.platforms:
            platform = device.get('platform_name', '').lower()
            if platform not in self.platforms:
                return False

        # Status filter
        if self.statuses:
            status = device.get('status', '').lower()
            if status not in self.statuses:
                return False

        # Group filter
        if self.groups:
            device_groups = device.get('groups', [])
            device_group_names = [host_groups.get(gid, '').lower() for gid in device_groups]
            # Check if any device group matches any filter group
            if not any(fg.lower() in dgn for fg in self.groups for dgn in device_group_names):
                return False

        # Stale devices filter
        if self.stale_days is not None:
            last_seen = device.get('last_seen', '')
            if last_seen:
                try:
                    last_seen_dt = datetime.strptime(last_seen, '%Y-%m-%dT%H:%M:%SZ')
                    if (datetime.utcnow() - last_seen_dt).days > self.stale_days:
                        return False
                except:
                    pass

        return True


def detect_anomalies(devices: List[Dict[str, Any]], policies: Dict[str, Dict[str, str]]) -> Dict[str, List[Dict[str, Any]]]:
    """
    Detect anomalies in device configurations.

    Returns:
        Dictionary with anomaly type as key and list of affected devices
    """
    anomalies = {
        'no_prevention_policy': [],
        'no_response_policy': [],
        'no_sensor_policy': [],
        'policy_not_applied': [],
        'no_host_group': [],
        'stale_devices': [],
    }

    for device in devices:
        device_id = device.get('device_id', '')
        hostname = device.get('hostname', '')
        last_seen = device.get('last_seen', '')

        # Check prevention policy
        prevention_policy_id = device.get('device_policies', {}).get('prevention', {}).get('policy_id')
        if not prevention_policy_id:
            anomalies['no_prevention_policy'].append({
                'device_id': device_id,
                'hostname': hostname,
                'issue': 'No prevention policy assigned'
            })
        else:
            prevention_applied = device.get('device_policies', {}).get('prevention', {}).get('applied', False)
            if not prevention_applied:
                anomalies['policy_not_applied'].append({
                    'device_id': device_id,
                    'hostname': hostname,
                    'issue': 'Prevention policy not applied'
                })

        # Check response policy
        response_policy_id = device.get('device_policies', {}).get('remote_response', {}).get('policy_id')
        if not response_policy_id:
            anomalies['no_response_policy'].append({
                'device_id': device_id,
                'hostname': hostname,
                'issue': 'No response policy assigned'
            })

        # Check sensor update policy
        sensor_policy_id = device.get('device_policies', {}).get('sensor_update', {}).get('policy_id')
        if not sensor_policy_id:
            anomalies['no_sensor_policy'].append({
                'device_id': device_id,
                'hostname': hostname,
                'issue': 'No sensor update policy assigned'
            })

        # Check host groups
        device_groups = device.get('groups', [])
        if not device_groups:
            anomalies['no_host_group'].append({
                'device_id': device_id,
                'hostname': hostname,
                'issue': 'Not in any host group'
            })

        # Check stale devices (not seen in 30+ days)
        if last_seen:
            try:
                last_seen_dt = datetime.strptime(last_seen, '%Y-%m-%dT%H:%M:%SZ')
                days_stale = (datetime.utcnow() - last_seen_dt).days
                if days_stale > 30:
                    anomalies['stale_devices'].append({
                        'device_id': device_id,
                        'hostname': hostname,
                        'issue': f'Not seen for {days_stale} days'
                    })
            except:
                pass

    return anomalies


def calculate_statistics(devices: List[Dict[str, Any]], host_groups: Dict[str, str],
                         policies: Dict[str, Dict[str, str]]) -> Dict[str, Any]:
    """
    Calculate statistics from device data.

    Returns:
        Dictionary with various statistics
    """
    stats = {
        'total_devices': len(devices),
        'by_platform': Counter(),
        'by_status': Counter(),
        'by_host_group': Counter(),
        'by_prevention_policy': Counter(),
        'by_response_policy': Counter(),
        'by_sensor_policy': Counter(),
    }

    for device in devices:
        # Platform
        platform = device.get('platform_name', 'Unknown')
        stats['by_platform'][platform] += 1

        # Status
        status = device.get('status', 'Unknown')
        stats['by_status'][status] += 1

        # Host groups
        device_groups = device.get('groups', [])
        if device_groups:
            for gid in device_groups:
                group_name = host_groups.get(gid, gid)
                stats['by_host_group'][group_name] += 1
        else:
            stats['by_host_group']['No group'] += 1

        # Prevention policy
        prev_policy_id = device.get('device_policies', {}).get('prevention', {}).get('policy_id')
        if prev_policy_id:
            policy_name = policies['prevention'].get(prev_policy_id, 'Unknown')
            stats['by_prevention_policy'][policy_name] += 1
        else:
            stats['by_prevention_policy']['No policy'] += 1

        # Response policy
        resp_policy_id = device.get('device_policies', {}).get('remote_response', {}).get('policy_id')
        if resp_policy_id:
            policy_name = policies['response'].get(resp_policy_id, 'Unknown')
            stats['by_response_policy'][policy_name] += 1
        else:
            stats['by_response_policy']['No policy'] += 1

        # Sensor policy
        sensor_policy_id = device.get('device_policies', {}).get('sensor_update', {}).get('policy_id')
        if sensor_policy_id:
            policy_name = policies['sensor_update'].get(sensor_policy_id, 'Unknown')
            stats['by_sensor_policy'][policy_name] += 1
        else:
            stats['by_sensor_policy']['No policy'] += 1

    return stats


def print_statistics(stats: Dict[str, Any], anomalies: Dict[str, List[Dict[str, Any]]]):
    """Print statistics and anomalies to console."""
    print_header("STATISTICS & ANOMALIES", width=80)
    print()

    # Total devices
    print(f"{Colors.HIGHLIGHT}Total Devices:{Colors.RESET} {stats['total_devices']}")
    print()

    # Platform distribution
    print(f"{Colors.INFO}Platform Distribution:{Colors.RESET}")
    total = stats['total_devices']
    for platform, count in stats['by_platform'].most_common():
        pct = (count / total * 100) if total > 0 else 0
        bar_length = int(pct / 2)
        bar = '█' * bar_length + '░' * (50 - bar_length)
        print(f"  {platform:20s} [{bar}] {count:4d} ({pct:5.1f}%)")
    print()

    # Status distribution
    print(f"{Colors.INFO}Status Distribution:{Colors.RESET}")
    for status, count in stats['by_status'].most_common():
        pct = (count / total * 100) if total > 0 else 0
        print(f"  {status:20s} {count:4d} ({pct:5.1f}%)")
    print()

    # Top host groups
    print(f"{Colors.INFO}Top 10 Host Groups:{Colors.RESET}")
    for group, count in stats['by_host_group'].most_common(10):
        print(f"  {group[:50]:50s} {count:4d}")
    print()

    # Anomalies
    print_header("ANOMALIES DETECTED", width=80)
    print()

    total_anomalies = sum(len(devices) for devices in anomalies.values())

    if total_anomalies == 0:
        print_success("No anomalies detected!")
    else:
        print_warning(f"Found {total_anomalies} anomaly/anomalies")
        print()

        anomaly_labels = {
            'no_prevention_policy': 'No Prevention Policy',
            'no_response_policy': 'No Response Policy',
            'no_sensor_policy': 'No Sensor Update Policy',
            'policy_not_applied': 'Policy Not Applied',
            'no_host_group': 'No Host Group',
            'stale_devices': 'Stale Devices (>30 days)',
        }

        for anomaly_type, label in anomaly_labels.items():
            count = len(anomalies[anomaly_type])
            if count > 0:
                print(f"  {Colors.WARNING}⚠{Colors.RESET} {label:30s} {count:4d} device(s)")

    print()


def get_all_cids(flight_control: FlightControl) -> Tuple[Dict[str, Any], List[Dict[str, Any]]]:
    """
    Get parent CID and all child CIDs in Flight Control.

    Args:
        flight_control: FlightControl API instance

    Returns:
        Tuple of (parent_info, children_list)
    """
    print_info("Retrieving Flight Control CID information...")

    # Get parent CID info
    parent_response = flight_control.query_children()

    if parent_response['status_code'] != 200:
        print_error("Failed to retrieve parent CID information")
        return None, []

    # Get parent CID from a sensor update policy (reliable source)
    from falconpy import SensorUpdatePolicies
    sensor_policies = SensorUpdatePolicies(auth_object=flight_control.auth_object)
    policies_response = sensor_policies.queryCombinedSensorUpdatePoliciesV2(limit=1)

    if policies_response['status_code'] == 200 and policies_response['body'].get('resources'):
        parent_cid = policies_response['body']['resources'][0].get('cid', 'Unknown')
    else:
        parent_cid = 'Unknown'

    parent_name = "Parent CID"

    parent_info = {
        'cid': parent_cid,
        'name': parent_name,
        'type': 'parent'
    }

    # Get children
    child_cids = parent_response['body'].get('resources', [])

    if not child_cids:
        print_warning("No child CIDs found")
        return parent_info, []

    # Get details for each child
    children = []
    if child_cids:
        details_response = flight_control.get_children(ids=child_cids)

        if details_response['status_code'] == 200:
            for child in details_response['body'].get('resources', []):
                children.append({
                    'cid': child.get('child_cid', 'Unknown'),
                    'name': child.get('name', 'Unknown'),
                    'type': 'child'
                })
        else:
            # Fallback: use CIDs without names
            for cid in child_cids:
                children.append({
                    'cid': cid,
                    'name': cid,
                    'type': 'child'
                })

    print_success(f"Found parent CID and {len(children)} child CID(s)")
    return parent_info, children


def select_cids_to_export(parent: Dict[str, Any], children: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    """
    Interactive selection of CIDs to export.

    Args:
        parent: Parent CID info
        children: List of child CIDs

    Returns:
        List of selected CIDs
    """
    print_header("SELECT CIDs TO EXPORT", width=80)

    all_cids = [parent] + children

    print()
    print(f"{Colors.INFO}Available CIDs:{Colors.RESET}\n")

    for i, cid in enumerate(all_cids, 1):
        cid_type = "PARENT" if cid['type'] == 'parent' else "CHILD"
        print(f"  {Colors.HIGHLIGHT}[{i}]{Colors.RESET} {Colors.BRIGHT}{cid['name']}{Colors.RESET} ({cid_type})")
        print(f"      {Colors.DIM}CID: {cid['cid']}{Colors.RESET}")
        print()

    print(f"{Colors.INFO}Selection options:{Colors.RESET}")
    print("  • Enter CID numbers separated by commas (e.g., 1,3,4)")
    print("  • Enter 'all' to select all CIDs")
    print("  • Enter 'children' to select all children only")
    print("  • Enter 'q' to quit")
    print()

    while True:
        selection = input(f"{Colors.HIGHLIGHT}Select CIDs to export: {Colors.RESET}").strip().lower()

        if selection == 'q':
            print_warning("Export cancelled")
            sys.exit(0)

        if selection == 'all':
            print_success(f"Selected all {len(all_cids)} CID(s)")
            return all_cids

        if selection == 'children':
            if not children:
                print_error("No child CIDs available")
                continue
            print_success(f"Selected {len(children)} child CID(s)")
            return children

        # Parse comma-separated numbers
        try:
            numbers = [int(n.strip()) for n in selection.split(',')]
            if not all(1 <= n <= len(all_cids) for n in numbers):
                print_error(f"Invalid selection. Enter numbers between 1 and {len(all_cids)}")
                continue

            selected = [all_cids[n-1] for n in numbers]
            print_success(f"Selected {len(selected)} CID(s):")
            for cid in selected:
                print(f"  • {cid['name']}")
            print()
            return selected

        except (ValueError, IndexError):
            print_error("Invalid input. Please enter numbers separated by commas, 'all', 'children', or 'q'")


def get_devices_for_cid(hosts: Hosts, cid_info: Dict[str, Any]) -> List[str]:
    """
    Get all device IDs for a specific CID.

    Args:
        hosts: Hosts API instance
        cid_info: CID information dict

    Returns:
        List of device IDs
    """
    import time
    import threading

    device_ids = []
    offset = None
    limit = 5000

    # Spinner animation
    spinner_running = True
    spinner_chars = ['⠋', '⠙', '⠹', '⠸', '⠼', '⠴', '⠦', '⠧', '⠇', '⠏']

    def spinner():
        idx = 0
        while spinner_running:
            sys.stdout.write(f'\r  {Colors.INFO}{spinner_chars[idx % len(spinner_chars)]}{Colors.RESET} Querying devices (this may take several minutes for large environments)...')
            sys.stdout.flush()
            idx += 1
            time.sleep(0.1)

    # Start spinner
    spinner_thread = threading.Thread(target=spinner, daemon=True)
    spinner_thread.start()

    try:
        while True:
            response = hosts.query_devices_by_filter(
                offset=offset,
                limit=limit
            )

            if response['status_code'] != 200:
                spinner_running = False
                time.sleep(0.2)  # Let spinner finish
                sys.stdout.write('\r' + ' ' * 100 + '\r')  # Clear line
                sys.stdout.flush()
                print_error(f"Failed to query devices for {cid_info['name']}")
                break

            batch_ids = response['body'].get('resources', [])

            # If no results in this batch, we're done
            if not batch_ids:
                break

            device_ids.extend(batch_ids)

            # Check pagination
            meta = response['body'].get('meta', {})
            pagination = meta.get('pagination', {})
            new_offset = pagination.get('offset')

            # If offset hasn't changed or is the same, we're done
            if not new_offset or new_offset == offset:
                break

            offset = new_offset
    finally:
        spinner_running = False
        time.sleep(0.2)  # Let spinner finish
        sys.stdout.write('\r' + ' ' * 100 + '\r')  # Clear line
        sys.stdout.flush()

    return device_ids


def get_device_details(hosts: Hosts, device_ids: List[str]) -> List[Dict[str, Any]]:
    """
    Get detailed information for devices.

    Args:
        hosts: Hosts API instance
        device_ids: List of device IDs

    Returns:
        List of device details
    """
    if not device_ids:
        return []

    import time
    import threading

    devices = []
    batch_size = 1000
    total_batches = (len(device_ids) + batch_size - 1) // batch_size

    # Progress indicator
    progress_running = True

    def progress_indicator():
        dots = 0
        while progress_running:
            sys.stdout.write(f'\r  {Colors.INFO}⏳{Colors.RESET} Retrieving device details (batch processing){"." * (dots % 4)}{" " * (3 - (dots % 4))}')
            sys.stdout.flush()
            dots += 1
            time.sleep(0.3)

    # Start progress indicator
    progress_thread = threading.Thread(target=progress_indicator, daemon=True)
    progress_thread.start()

    try:
        for i in range(0, len(device_ids), batch_size):
            batch = device_ids[i:i + batch_size]

            response = hosts.get_device_details(ids=batch)

            if response['status_code'] == 200:
                devices.extend(response['body'].get('resources', []))
            else:
                # Don't stop progress for warnings
                pass
    finally:
        progress_running = False
        time.sleep(0.4)  # Let progress finish
        sys.stdout.write('\r' + ' ' * 100 + '\r')  # Clear line
        sys.stdout.flush()

    return devices


def get_host_groups(host_group: HostGroup, devices: List[Dict[str, Any]]) -> Dict[str, str]:
    """
    Get all host groups and create ID to name mapping.

    Args:
        host_group: HostGroup API instance
        devices: List of devices (to collect all group IDs)

    Returns:
        Dictionary mapping group ID to group name
    """
    groups = {}

    # Collect all unique group IDs from devices
    all_group_ids = set()
    for device in devices:
        device_groups = device.get('groups', [])
        all_group_ids.update(device_groups)

    if not all_group_ids:
        return {}

    # Try to get details for all group IDs in batches
    group_ids_list = list(all_group_ids)
    batch_size = 100

    for i in range(0, len(group_ids_list), batch_size):
        batch = group_ids_list[i:i + batch_size]

        try:
            details_response = host_group.get_host_groups(ids=batch)

            if details_response['status_code'] == 200:
                for group in details_response['body'].get('resources', []):
                    groups[group['id']] = group.get('name', group['id'])
        except:
            # If batch fails, try individually
            for group_id in batch:
                try:
                    details_response = host_group.get_host_groups(ids=[group_id])
                    if details_response['status_code'] == 200:
                        for group in details_response['body'].get('resources', []):
                            groups[group['id']] = group.get('name', group['id'])
                except:
                    pass

    # For any group IDs we couldn't resolve, use the ID itself
    for group_id in all_group_ids:
        if group_id not in groups:
            groups[group_id] = group_id  # Use ID as fallback

    return groups


def get_policies(prevention: PreventionPolicy, response_policy: ResponsePolicies,
                 sensor_update: SensorUpdatePolicies) -> Dict[str, Dict[str, str]]:
    """
    Get all policies and create ID to name mappings.

    Args:
        prevention: PreventionPolicy API instance
        response_policy: ResponsePolicies API instance
        sensor_update: SensorUpdatePolicies API instance

    Returns:
        Dictionary with policy type as key and ID->name mappings as values
    """
    policies = {
        'prevention': {},
        'response': {},
        'sensor_update': {}
    }

    # Prevention policies
    prev_response = prevention.queryCombinedPreventionPolicies()
    if prev_response['status_code'] == 200:
        for policy in prev_response['body'].get('resources', []):
            policies['prevention'][policy['id']] = policy.get('name', 'Unknown')

    # Response policies
    resp_response = response_policy.queryCombinedRTResponsePolicies()
    if resp_response['status_code'] == 200:
        for policy in resp_response['body'].get('resources', []):
            policies['response'][policy['id']] = policy.get('name', 'Unknown')

    # Sensor update policies
    sensor_response = sensor_update.queryCombinedSensorUpdatePoliciesV2()
    if sensor_response['status_code'] == 200:
        for policy in sensor_response['body'].get('resources', []):
            policies['sensor_update'][policy['id']] = policy.get('name', 'Unknown')

    return policies


def export_cid_to_csv(cid_info: Dict[str, Any], devices: List[Dict[str, Any]],
                      host_groups: Dict[str, str], policies: Dict[str, Dict[str, str]],
                      filters: Optional[DeviceFilters] = None) -> List[Dict[str, Any]]:
    """
    Convert device data to CSV rows.

    Args:
        cid_info: CID information
        devices: List of device details
        host_groups: Host group ID to name mapping
        policies: Policy mappings
        filters: Optional device filters

    Returns:
        List of CSV row dictionaries
    """
    rows = []

    for device in devices:
        # Apply filters if provided
        if filters and not filters.should_include(device, host_groups):
            continue

        # Get host groups
        device_groups = device.get('groups', [])
        group_names = [host_groups.get(gid, gid) for gid in device_groups]

        # Get prevention policy
        prevention_policy_id = device.get('device_policies', {}).get('prevention', {}).get('policy_id')
        prevention_applied = device.get('device_policies', {}).get('prevention', {}).get('applied', False)
        prevention_policy_name = policies['prevention'].get(prevention_policy_id, 'None') if prevention_policy_id else 'None'
        prevention_status = 'Applied' if prevention_applied else 'Assigned'

        # Get response policy
        response_policy_id = device.get('device_policies', {}).get('remote_response', {}).get('policy_id')
        response_applied = device.get('device_policies', {}).get('remote_response', {}).get('applied', False)
        response_policy_name = policies['response'].get(response_policy_id, 'None') if response_policy_id else 'None'
        response_status = 'Applied' if response_applied else 'Assigned'

        # Get sensor update policy
        sensor_policy_id = device.get('device_policies', {}).get('sensor_update', {}).get('policy_id')
        sensor_applied = device.get('device_policies', {}).get('sensor_update', {}).get('applied', False)
        sensor_policy_name = policies['sensor_update'].get(sensor_policy_id, 'None') if sensor_policy_id else 'None'
        sensor_status = 'Applied' if sensor_applied else 'Assigned'

        row = {
            'CID Name': cid_info['name'],
            'CID': cid_info['cid'],
            'CID Type': cid_info['type'].upper(),
            'Device ID': device.get('device_id', ''),
            'Hostname': device.get('hostname', ''),
            'OS Version': device.get('os_version', ''),
            'Platform': device.get('platform_name', ''),
            'Last Seen': device.get('last_seen', ''),
            'Status': device.get('status', ''),
            'Host Groups': ', '.join(group_names) if group_names else 'None',
            'Prevention Policy': prevention_policy_name,
            'Prevention Status': prevention_status if prevention_policy_id else 'None',
            'Response Policy': response_policy_name,
            'Response Status': response_status if response_policy_id else 'None',
            'Sensor Update Policy': sensor_policy_name,
            'Sensor Update Status': sensor_status if sensor_policy_id else 'None',
            'Agent Version': device.get('agent_version', ''),
            'Service Provider': device.get('service_provider', ''),
            'Service Provider Account ID': device.get('service_provider_account_id', ''),
        }

        rows.append(row)

    return rows


def export_to_excel(output_file: str, all_data: Dict[str, List[Dict[str, Any]]],
                    all_stats: Dict[str, Dict[str, Any]], all_anomalies: Dict[str, Dict[str, List[Dict[str, Any]]]]):
    """
    Export data to Excel with multiple sheets and formatting.

    Args:
        output_file: Output file path
        all_data: Dictionary with CID name as key and device rows as value
        all_stats: Dictionary with CID name as key and statistics as value
        all_anomalies: Dictionary with CID name as key and anomalies as value
    """
    if not EXCEL_AVAILABLE:
        print_warning("Excel export not available (openpyxl not installed)")
        return

    wb = Workbook()

    # Styles
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    warning_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    error_fill = PatternFill(start_color="F4CCCC", end_color="F4CCCC", fill_type="solid")
    success_fill = PatternFill(start_color="D9EAD3", end_color="D9EAD3", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Remove default sheet
    if 'Sheet' in wb.sheetnames:
        del wb['Sheet']

    # Summary sheet
    ws_summary = wb.create_sheet("Summary", 0)
    ws_summary['A1'] = "Devices & Policies Export Summary"
    ws_summary['A1'].font = Font(size=16, bold=True)
    ws_summary.merge_cells('A1:D1')

    ws_summary['A3'] = "Generated:"
    ws_summary['B3'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')

    row = 5
    ws_summary[f'A{row}'] = "CID Name"
    ws_summary[f'B{row}'] = "Total Devices"
    ws_summary[f'C{row}'] = "Anomalies"
    ws_summary[f'D{row}'] = "Status"

    for col in ['A', 'B', 'C', 'D']:
        ws_summary[f'{col}{row}'].font = header_font
        ws_summary[f'{col}{row}'].fill = header_fill

    for cid_name in all_data.keys():
        row += 1
        device_count = len(all_data[cid_name])
        anomaly_count = sum(len(devices) for devices in all_anomalies.get(cid_name, {}).values())

        ws_summary[f'A{row}'] = cid_name
        ws_summary[f'B{row}'] = device_count
        ws_summary[f'C{row}'] = anomaly_count

        status_cell = ws_summary[f'D{row}']
        if anomaly_count == 0:
            status_cell.value = "✓ Clean"
            status_cell.fill = success_fill
        elif anomaly_count < device_count * 0.1:
            status_cell.value = "⚠ Minor issues"
            status_cell.fill = warning_fill
        else:
            status_cell.value = "✗ Issues"
            status_cell.fill = error_fill

    ws_summary.column_dimensions['A'].width = 40
    ws_summary.column_dimensions['B'].width = 15
    ws_summary.column_dimensions['C'].width = 15
    ws_summary.column_dimensions['D'].width = 20

    # Device sheets (one per CID)
    fieldnames = [
        'CID Name', 'CID', 'CID Type', 'Device ID', 'Hostname', 'OS Version', 'Platform',
        'Last Seen', 'Status', 'Host Groups', 'Prevention Policy', 'Prevention Status',
        'Response Policy', 'Response Status', 'Sensor Update Policy', 'Sensor Update Status',
        'Agent Version', 'Service Provider', 'Service Provider Account ID'
    ]

    for cid_name, rows in all_data.items():
        if not rows:
            continue

        # Sanitize sheet name
        sheet_name = cid_name[:31].replace('/', '-').replace('\\', '-').replace(':', '-')
        ws = wb.create_sheet(sheet_name)

        # Headers
        for col_idx, field in enumerate(fieldnames, 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.value = field
            cell.font = header_font
            cell.fill = header_fill
            cell.border = thin_border

        # Data
        for row_idx, row_data in enumerate(rows, 2):
            for col_idx, field in enumerate(fieldnames, 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.value = row_data.get(field, '')
                cell.border = thin_border

                # Color coding for status
                if field == 'Prevention Status' or field == 'Response Status' or field == 'Sensor Update Status':
                    if cell.value == 'Applied':
                        cell.fill = success_fill
                    elif cell.value == 'Assigned':
                        cell.fill = warning_fill
                    elif cell.value == 'None':
                        cell.fill = error_fill

        # Auto-filter
        ws.auto_filter.ref = ws.dimensions

        # Freeze panes (first row)
        ws.freeze_panes = 'A2'

        # Auto-adjust column widths
        for col_idx, field in enumerate(fieldnames, 1):
            max_length = len(field)
            for row_idx in range(2, min(len(rows) + 2, 100)):  # Check first 100 rows
                cell_value = ws.cell(row=row_idx, column=col_idx).value
                if cell_value:
                    max_length = max(max_length, len(str(cell_value)))
            ws.column_dimensions[get_column_letter(col_idx)].width = min(max_length + 2, 50)

    # Anomalies sheet
    if any(all_anomalies.values()):
        ws_anomalies = wb.create_sheet("Anomalies")

        ws_anomalies['A1'] = "Detected Anomalies"
        ws_anomalies['A1'].font = Font(size=14, bold=True)
        ws_anomalies.merge_cells('A1:D1')

        row = 3
        ws_anomalies[f'A{row}'] = "CID"
        ws_anomalies[f'B{row}'] = "Device ID"
        ws_anomalies[f'C{row}'] = "Hostname"
        ws_anomalies[f'D{row}'] = "Issue"

        for col in ['A', 'B', 'C', 'D']:
            ws_anomalies[f'{col}{row}'].font = header_font
            ws_anomalies[f'{col}{row}'].fill = header_fill

        for cid_name, anomalies in all_anomalies.items():
            for anomaly_type, devices in anomalies.items():
                for device_info in devices:
                    row += 1
                    ws_anomalies[f'A{row}'] = cid_name
                    ws_anomalies[f'B{row}'] = device_info.get('device_id', '')
                    ws_anomalies[f'C{row}'] = device_info.get('hostname', '')
                    ws_anomalies[f'D{row}'] = device_info.get('issue', '')
                    ws_anomalies[f'D{row}'].fill = warning_fill

        ws_anomalies.column_dimensions['A'].width = 30
        ws_anomalies.column_dimensions['B'].width = 40
        ws_anomalies.column_dimensions['C'].width = 30
        ws_anomalies.column_dimensions['D'].width = 40

    # Save workbook
    wb.save(output_file)
    print_success(f"Excel file created: {output_file}")


def main():
    """Main function."""
    parser = argparse.ArgumentParser(
        description='Export CrowdStrike devices, groups, and policies to CSV/Excel'
    )
    parser.add_argument(
        '--config',
        type=str,
        help='Path to credentials config file'
    )
    parser.add_argument(
        '--client-id',
        type=str,
        help='Falcon API Client ID'
    )
    parser.add_argument(
        '--client-secret',
        type=str,
        help='Falcon API Client Secret'
    )
    parser.add_argument(
        '--base-url',
        type=str,
        default='https://api.crowdstrike.com',
        help='Falcon API base URL'
    )
    parser.add_argument(
        '--output',
        type=str,
        help='Output file path (default: devices_export_TIMESTAMP with auto-detected format)'
    )
    parser.add_argument(
        '--format',
        type=str,
        choices=['csv', 'excel', 'both'],
        default='excel',
        help='Output format (default: excel)'
    )
    parser.add_argument(
        '--non-interactive',
        action='store_true',
        help='Export all CIDs without prompting'
    )
    parser.add_argument(
        '--filter-platform',
        type=str,
        help='Filter by platform (comma-separated: Windows,Linux,Mac)'
    )
    parser.add_argument(
        '--filter-status',
        type=str,
        help='Filter by status (comma-separated: normal,containment)'
    )
    parser.add_argument(
        '--filter-groups',
        type=str,
        help='Filter by host groups (comma-separated, partial match)'
    )
    parser.add_argument(
        '--stale-threshold',
        type=int,
        help='Filter out devices not seen in X days'
    )

    args = parser.parse_args()

    print_jg_logo()
    print_header("FLIGHT CONTROL - DEVICES & POLICIES EXPORT", width=80, color=Colors.SUCCESS)

    # Get credentials
    client_id, client_secret, base_url, source = get_credentials_smart(
        config_path=args.config,
        client_id=args.client_id,
        client_secret=args.client_secret,
        base_url=args.base_url
    )

    if not client_id or not client_secret:
        print_error("No credentials provided!")
        print()
        print("Please provide credentials via one of these methods:")
        print("  1. Config file: --config config/credentials.json")
        print("  2. CLI args: --client-id <id> --client-secret <secret>")
        print("  3. Environment variables: FALCON_CLIENT_ID, FALCON_CLIENT_SECRET")
        sys.exit(1)

    from utils.formatting import print_credentials_source
    print_credentials_source(source)

    # Authenticate
    print_info("Authenticating to Falcon API...")

    try:
        flight_control = FlightControl(
            client_id=client_id,
            client_secret=client_secret,
            base_url=base_url
        )
        print_success("Authentication successful!")
    except Exception as e:
        print_error(f"Authentication failed: {str(e)}")
        sys.exit(1)

    # Get all CIDs
    parent, children = get_all_cids(flight_control)

    if not parent:
        print_error("Failed to retrieve CID information")
        sys.exit(1)

    # Select CIDs
    if args.non_interactive:
        selected_cids = [parent] + children
        print_info(f"Non-interactive mode: exporting all {len(selected_cids)} CID(s)")
    else:
        selected_cids = select_cids_to_export(parent, children)

    if not selected_cids:
        print_warning("No CIDs selected")
        sys.exit(0)

    # Create device filters
    filters = None
    if any([args.filter_platform, args.filter_status, args.filter_groups, args.stale_threshold]):
        platforms = args.filter_platform.split(',') if args.filter_platform else None
        statuses = args.filter_status.split(',') if args.filter_status else None
        groups = args.filter_groups.split(',') if args.filter_groups else None

        filters = DeviceFilters(
            platforms=platforms,
            statuses=statuses,
            groups=groups,
            stale_days=args.stale_threshold
        )

        print()
        print_info("Active filters:")
        if platforms:
            print(f"  • Platforms: {', '.join(platforms)}")
        if statuses:
            print(f"  • Statuses: {', '.join(statuses)}")
        if groups:
            print(f"  • Groups: {', '.join(groups)}")
        if args.stale_threshold:
            print(f"  • Exclude devices not seen in {args.stale_threshold} days")

    # Prepare output files
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')

    if args.output:
        # Use provided output path
        base_output = args.output
        if base_output.endswith('.csv'):
            csv_output = base_output
            excel_output = base_output.replace('.csv', '.xlsx')
        elif base_output.endswith('.xlsx'):
            excel_output = base_output
            csv_output = base_output.replace('.xlsx', '.csv')
        else:
            csv_output = f"{base_output}.csv"
            excel_output = f"{base_output}.xlsx"
    else:
        # Default filenames
        csv_output = f"devices_export_{timestamp}.csv"
        excel_output = f"devices_export_{timestamp}.xlsx"

    # Ensure output directory exists
    for output in [csv_output, excel_output]:
        output_dir = os.path.dirname(output)
        if output_dir:
            os.makedirs(output_dir, exist_ok=True)

    print_header("EXPORTING DEVICE DATA", width=80)

    all_rows = []
    all_data = {}  # For Excel export by CID
    all_stats = {}
    all_anomalies = {}
    total_devices = 0
    total_filtered = 0

    for idx, cid_info in enumerate(selected_cids, 1):
        print()
        print(f"{Colors.HIGHLIGHT}▶ Processing: {cid_info['name']} ({cid_info['type'].upper()}){Colors.RESET}")

        # Authenticate with specific CID if it's a child
        if cid_info['type'] == 'child':
            hosts = Hosts(
                client_id=client_id,
                client_secret=client_secret,
                base_url=base_url,
                member_cid=cid_info['cid']
            )
            host_group = HostGroup(
                client_id=client_id,
                client_secret=client_secret,
                base_url=base_url,
                member_cid=cid_info['cid']
            )
            prevention = PreventionPolicy(
                client_id=client_id,
                client_secret=client_secret,
                base_url=base_url,
                member_cid=cid_info['cid']
            )
            response_policy = ResponsePolicies(
                client_id=client_id,
                client_secret=client_secret,
                base_url=base_url,
                member_cid=cid_info['cid']
            )
            sensor_update = SensorUpdatePolicies(
                client_id=client_id,
                client_secret=client_secret,
                base_url=base_url,
                member_cid=cid_info['cid']
            )
        else:
            hosts = Hosts(
                client_id=client_id,
                client_secret=client_secret,
                base_url=base_url
            )
            host_group = HostGroup(
                client_id=client_id,
                client_secret=client_secret,
                base_url=base_url
            )
            prevention = PreventionPolicy(
                client_id=client_id,
                client_secret=client_secret,
                base_url=base_url
            )
            response_policy = ResponsePolicies(
                client_id=client_id,
                client_secret=client_secret,
                base_url=base_url
            )
            sensor_update = SensorUpdatePolicies(
                client_id=client_id,
                client_secret=client_secret,
                base_url=base_url
            )

        # Get devices
        print_info("  Querying devices...")
        device_ids = get_devices_for_cid(hosts, cid_info)
        print_success(f"  Found {len(device_ids)} device(s)")

        if not device_ids:
            print_warning(f"  No devices found for {cid_info['name']}")
            continue

        # Get device details
        print_info("  Retrieving device details...")
        devices = get_device_details(hosts, device_ids)
        print_success(f"  Retrieved details for {len(devices)} device(s)")

        # Get host groups (pass devices to collect all group IDs)
        print_info("  Loading host groups...")
        host_groups = get_host_groups(host_group, devices)
        print_success(f"  Loaded {len(host_groups)} host group(s)")

        # Get policies
        print_info("  Loading policies...")
        policies = get_policies(prevention, response_policy, sensor_update)
        policy_count = len(policies['prevention']) + len(policies['response']) + len(policies['sensor_update'])
        print_success(f"  Loaded {policy_count} policie(s)")

        # Convert to CSV rows
        print_info("  Converting to CSV format...")
        rows = export_cid_to_csv(cid_info, devices, host_groups, policies, filters)
        all_rows.extend(rows)
        all_data[cid_info['name']] = rows
        total_devices += len(devices)
        total_filtered += len(rows)
        print_success(f"  Processed {len(rows)} device(s) (filtered from {len(devices)})")

        # Calculate statistics
        print_info("  Calculating statistics...")
        stats = calculate_statistics(rows, host_groups, policies)
        all_stats[cid_info['name']] = stats

        # Detect anomalies
        print_info("  Detecting anomalies...")
        anomalies = detect_anomalies(rows, policies)
        all_anomalies[cid_info['name']] = anomalies

        anomaly_count = sum(len(devices) for devices in anomalies.values())
        if anomaly_count > 0:
            print_warning(f"  Found {anomaly_count} anomaly/anomalies")
        else:
            print_success(f"  No anomalies detected")

        # Progress
        print_progress(idx, len(selected_cids), prefix="Overall progress", suffix=f"({cid_info['name'][:30]})")

    # Write output files
    print()

    if not all_rows:
        print_warning("No data to export (all devices filtered out)")
    else:
        fieldnames = [
            'CID Name', 'CID', 'CID Type', 'Device ID', 'Hostname', 'OS Version', 'Platform',
            'Last Seen', 'Status', 'Host Groups', 'Prevention Policy', 'Prevention Status',
            'Response Policy', 'Response Status', 'Sensor Update Policy', 'Sensor Update Status',
            'Agent Version', 'Service Provider', 'Service Provider Account ID'
        ]

        # CSV export
        if args.format in ['csv', 'both']:
            print_info(f"Writing CSV file: {csv_output}")
            with open(csv_output, 'w', newline='', encoding='utf-8') as f:
                writer = csv.DictWriter(f, fieldnames=fieldnames)
                writer.writeheader()
                writer.writerows(all_rows)
            print_success(f"CSV file created successfully!")

        # Excel export
        if args.format in ['excel', 'both']:
            print_info(f"Writing Excel file: {excel_output}")
            export_to_excel(excel_output, all_data, all_stats, all_anomalies)

    # Print overall statistics
    print()
    if total_filtered > 0:
        # Merge stats from all CIDs
        merged_stats = {
            'total_devices': total_filtered,
            'by_platform': Counter(),
            'by_status': Counter(),
            'by_host_group': Counter(),
            'by_prevention_policy': Counter(),
            'by_response_policy': Counter(),
            'by_sensor_policy': Counter(),
        }

        for stats in all_stats.values():
            for key in ['by_platform', 'by_status', 'by_host_group', 'by_prevention_policy',
                       'by_response_policy', 'by_sensor_policy']:
                merged_stats[key].update(stats[key])

        # Merge anomalies from all CIDs
        merged_anomalies = {
            'no_prevention_policy': [],
            'no_response_policy': [],
            'no_sensor_policy': [],
            'policy_not_applied': [],
            'no_host_group': [],
            'stale_devices': [],
        }

        for anomalies in all_anomalies.values():
            for key in merged_anomalies.keys():
                merged_anomalies[key].extend(anomalies.get(key, []))

        print_statistics(merged_stats, merged_anomalies)

    # Summary
    summary_data = {
        'CIDs processed': len(selected_cids),
        'Total devices queried': total_devices,
        'Devices exported': total_filtered,
    }

    if filters:
        summary_data['Devices filtered out'] = total_devices - total_filtered

    if args.format in ['csv', 'both']:
        summary_data['CSV output'] = csv_output

    if args.format in ['excel', 'both']:
        summary_data['Excel output'] = excel_output

    print_summary_box(
        "EXPORT COMPLETE",
        summary_data,
        width=80
    )


if __name__ == '__main__':
    try:
        main()
    except KeyboardInterrupt:
        print()
        print_warning("Export interrupted by user")
        sys.exit(1)
    except Exception as e:
        print()
        print_error(f"Unexpected error: {str(e)}")
        import traceback
        traceback.print_exc()
        sys.exit(1)
